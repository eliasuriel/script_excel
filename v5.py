import pandas as pd
import numpy as np
import math
import os
import csv
import openpyxl
from openpyxl.styles import PatternFill

import tkinter as tk
from tkinter import filedialog


#print("Asking for input file")
#root = tk.Tk()
#root.withdraw()

#file_path = filedialog.askopenfilename()

#print("Reading File")
#df = pd.read_excel("Script_Input.xlsx") 

# Reemplaza 'nombre_del_archivo.xlsx' con el nombre de tu archivo Excel
excel_file = pd.ExcelFile('C:\\Users\OJA5GA\Documents\VS_Code\V1\Script_Input.xlsx')

# Reemplaza 'nombre_de_la_hoja' con el nombre de la hoja en tu archivo Excel
df = excel_file.parse('Sheet1')

# Reemplaza los valores nulos en la columna 2 con ceros
df['Sep'] = df['Sep'].fillna(0)
columna2 = df['Sep']
columna1 = df['ResourceName']
df['EGB_Group'] = '' 

Productivity1 = 0
Productivity2 = 0
horas = []
horas2 = []
EGB_Group = []
Nombres_Tareas = []
cont = 0
conditional = 0


WORK_DAYS = int(input("How many work days are in this report?: ") )
print(WORK_DAYS)

for index, row in df.iterrows():
    columna_tareas = row['ResourceName']
    columna_horas = row['Sep']
    #print(columna_tareas)
    #print(columna_horas)

    if 'EGB' in columna_tareas:
        cont = cont + 1
        Nombres_Tareas.append(columna_tareas)
        conditional = 1

        if Productivity1 !=0:
            Productivity1 = Productivity1/148
            horas.append(Productivity1)
            Productivity1 = 0
        else:
            if(cont >= 2):
                horas.append(Productivity1)
        if  Productivity2 !=0:
            Productivity2 = Productivity2/148
            horas2.append(Productivity2)
            Productivity2= 0

        if 'EGB3' in columna_tareas:
            EGB_Group.append('EGB3')
        elif 'EGB8' in columna_tareas:
            EGB_Group.append('EGB8')
        elif 'EGB9' in columna_tareas:
            EGB_Group.append('EGB9')
        elif 'EGB10' in columna_tareas:
            EGB_Group.append('EGB10')
        elif 'EGB11' in columna_tareas:
            EGB_Group.append('EGB11')
        elif 'EGB12' in columna_tareas:
            EGB_Group.append('EGB12')
        else:
            EGB_Group.append(' ')

    elif columna_tareas.startswith("   P_") and conditional == 1 and columna_horas != 0:
        Nombres_Tareas.append(columna_tareas)
        EGB_Group.append(' ')
        Productivity1 = Productivity1 + columna_horas
        Productivity2 = Productivity2 + columna_horas
    
    elif conditional == 1  and columna_horas != 0:
        Nombres_Tareas.append(columna_tareas)
        EGB_Group.append(' ')
    
    elif not 'EGB' in columna_tareas and (columna_tareas.endswith('-MS)') or columna_tareas.endswith('-MX)') or columna_tareas.endswith('-SX)')):
        conditional = 0
    




###########################################################
#                   Archivo Output
# #########################################################
    
  

archivo = "Monthly_Output.xlsx"
directorio = "C:\\Users\OJA5GA\Documents\VS_Code\V1"
mes = "Sep"
HOURS_PER_DAY = 8.25

ruta_completa = os.path.join(directorio, archivo)
print(ruta_completa)

# Verificar si el archivo existe

if not os.path.exists(ruta_completa):
    workbook = openpyxl.Workbook()      #Crear un nuevo libro
    hoja = workbook.active              #Crear una nueva hoja
    hoja.title = mes                    #Título de la hoja = mes
    workbook.save(ruta_completa)        #Guardar el libro

    print(f"Se ha creado el archivo {archivo} en: {ruta_completa}")

else:
    print(f"El archivo {archivo} ya existe en: {ruta_completa}")
    workbook = openpyxl.load_workbook(ruta_completa)
    hoja = workbook.active              #Crear una nueva hoja
    hoja.title = mes                    #Título de la hoja = mes
    workbook.save(ruta_completa)        #Guardar el libro

#Imprimir encabezados
hoja['A1'] = 'Resource Name'
hoja['B1'] = 'Hours'
hoja['C1'] = 'Productivity'
hoja['D1'] = 'Comments'
hoja['E1'] = 'EGB Group'
hoja['G1'] = 'Days'
hoja['G2'] = WORK_DAYS
hoja['H1'] = 'Hr x day'
hoja['H2'] = HOURS_PER_DAY
hoja['I1'] = 'Total hrs'
hoja['I2'] = HOURS_PER_DAY*WORK_DAYS



#Resultados
hoja['G5'] = 'Productivity Sum(%)'
hoja['G6'] = '=SUMA(C:C)'
hoja['H5'] = 'Qty People'
hoja['H6'] = '=CONTAR(C:C)'
hoja['I5'] = '% total'
hoja['I6'] = '=G6/H6'


#Imprimir nombres y tareas
for filas, contenido in enumerate(Nombres_Tareas, start=1):
    hoja.cell(row=filas+1, column=1, value=contenido)

#Imprimir EGB group
for filas, contenido in enumerate(EGB_Group, start=1):
    hoja.cell(row=filas+1, column=5, value=contenido)






# Itera a través de todas las columnas en la hoja para ajustar su ancho
for columna in hoja.columns:
    longitud_maxima = 0
    columna_letra = columna[0].column_letter  # Obtiene la letra de la columna
    for celda in columna:
        if celda.value:
            # Calcula la longitud máxima del contenido en la columna
            longitud_celda = len(str(celda.value))
            if longitud_celda > longitud_maxima:
                longitud_maxima = longitud_celda

    # Establece el ancho de la columna para ajustarlo al contenido más largo
    hoja.column_dimensions[columna_letra].width = longitud_maxima + 5

workbook.save(ruta_completa)







#print(Nombres_Tareas)  
#df = df[df['Sep'] != 0]


        

#print(cont)
#print(horas)
#print("OTRA \n" )
#print(horas2)
#print(df)