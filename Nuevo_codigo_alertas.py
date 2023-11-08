import pandas as pd
import numpy as np
import math
import os
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import NamedStyle
from openpyxl.worksheet.table import Table, TableStyleInfo
 
import tkinter as tk
from tkinter import filedialog
import easygui
import sys


#ccbox returns a boolean value 
def result_button(output):
    if output:
        #Continue
        pass 
    else:
        #Exit
        msg = easygui.msgbox("You decided to leave the program...", "Exit")
        sys.exit(0)
 
easygui.buttonbox("Choose your file to be proccessed (Excel File)",choices=["NEXT"])
#print("Asking for input file")
root = tk.Tk()
root.withdraw()
 
file_path = filedialog.askopenfilename(filetypes=[('Excel files', '*.xlsx')])

# button names
choices = ["Continue", "Exit"]

# Reemplaza 'nombre_del_archivo.xlsx' con el nombre de tu archivo Excel
excel_file = pd.ExcelFile(file_path)


#easygui.msgbox("The file was processed successfully")

result_button(easygui.ccbox("The file was processed successfully", "", choices))

#en caso de equivocarse de nombre de hoja lanza una alerta y vuelve a preguntar
while True:
    hojas = easygui.enterbox("In which sheet do you want to work? ")
   
    if hojas in excel_file.sheet_names:
        break
    else:
        easygui.msgbox(f"Sheet '{hojas}' not found in the Excel file. Please enter a valid sheet name.", "Sheet Not Found", choices)
 
# Reemplaza 'nombre_de_la_hoja' con el nombre de la hoja en tu archivo Excel
df = excel_file.parse(hojas)
 
# Reemplaza los valores de las columnas con nombres válidos
 
#en caso de equivocarse de nombre de columna lanza una alerta y vuelve a preguntar
while True:
    nombrecolumna1 = easygui.enterbox("Put the name of the first column you want to work on (coworkers, name of the jobs) ")
    if nombrecolumna1 in df.columns:
        break
    else:
        easygui.msgbox(f"Column '{nombrecolumna1}' not found in the selected sheet. Please enter a valid column name.", "Column Not Found")
 
while True:
    nombrecolumna2 = easygui.enterbox("Put the name of the second column you want to work on (hours) ")
    if nombrecolumna2 in df.columns:
        break
    else:
        easygui.msgbox(f"Column '{nombrecolumna2}' not found in the selected sheet. Please enter a valid column name.", "Column Not Found")
 
# Reemplaza los valores nulos en la columna 2 con ceros
df[nombrecolumna2] = df[nombrecolumna2].fillna(0)
columna2 = df[nombrecolumna2]
columna1 = df[nombrecolumna1]


#button = "Let's Go"
#easygui.msgbox("The names of the sheet and columns are correct",title=" ",ok_button=button)

result_button(easygui.ccbox("The names of the sheet and columns are correct", "", choices))


df['EGB_Group'] = ''
 
 
Nombres_Tareas = []
hours = []
Productivity = []
EGB_Group = []
Groups = ["EGB8","EGB9", "EGB10", "EGB11", "EGB12"]
People_per_group = []
 
cont = 0
conditional = 0
Productivity_abs = 0
Productivity_rel = 0
Productivity_sum = 0
Var_productivity = 0
cont_rel = 0
 
Productivity_abs_EGB8 = 0
Productivity_abs_EGB9 = 0
Productivity_abs_EGB10 = 0
Productivity_abs_EGB11 = 0
Productivity_abs_EGB12 = 0
Productivity_rel_EGB8 = 0
Productivity_rel_EGB9 = 0
Productivity_rel_EGB10 = 0
Productivity_rel_EGB11 = 0
Productivity_rel_EGB12 = 0
last_group = " "
cont_EGB8 = 0
cont_EGB9 = 0
cont_EGB10 = 0
cont_EGB11 = 0
cont_EGB12 = 0
 
cont_relEGB8 = 0
cont_relEGB9 = 0
cont_relEGB10 = 0
cont_relEGB11 = 0
cont_relEGB12 = 0
 
cont_grupo =0
 
cont_grupo_EGB8 =0
cont_grupo_EGB9 =0
cont_grupo_EGB10 =0
cont_grupo_EGB11 =0
cont_grupo_EGB12 =0
 
Var_productivity_EGB8 = 0
Var_productivity_EGB9 = 0
Var_productivity_EGB10 = 0
Var_productivity_EGB11 = 0
Var_productivity_EGB12 = 0
 
 
while True:
    WORK_DAYS = easygui.enterbox("How many work days are in this report?: ")
    try:
        WORK_DAYS = int(WORK_DAYS)
        if 1 <= WORK_DAYS <= 28:
            break
        else:
            easygui.msgbox("Invalid input. Please enter a number between 1 and 28.", "Invalid Input")
    except ValueError:
        easygui.msgbox("Invalid input. Please enter a valid number.", "Invalid Input")
 
#easygui.msgbox(WORK_DAYS, title="Días de trabajo", ok_button="NICE")
result_button(easygui.ccbox(WORK_DAYS, "", choices))
 
for index, row in df.iterrows():
    columna_tareas = row[nombrecolumna1]
    columna_horas = row[nombrecolumna2]
   
 
    if 'EGB' in columna_tareas and 'EGB3' not in columna_tareas:
       
        cont = cont + 1
        Nombres_Tareas.append(columna_tareas)
        hours.append(columna_horas)
        conditional = 1
        if columna_horas == 0:
            cont_grupo = cont_grupo + 1
       
   
        if cont > 1:
            Productivity.append(Var_productivity/148)
           
 
            if Var_productivity != 0:
                cont_rel = cont_rel + 1
 
            if Var_productivity_EGB8 !=0:
                cont_relEGB8 = cont_relEGB8 + 1
            if Var_productivity_EGB9 !=0:
                cont_relEGB9 = cont_relEGB9 + 1
            if Var_productivity_EGB10 !=0:
                cont_relEGB10 = cont_relEGB10 + 1
            if Var_productivity_EGB11 !=0:
                cont_relEGB11 = cont_relEGB11 + 1
            if Var_productivity_EGB12 !=0:
                cont_relEGB12 = cont_relEGB12 + 1
 
            Var_productivity = 0
            Var_productivity_EGB8 = 0
            Var_productivity_EGB9 = 0
            Var_productivity_EGB10 = 0
            Var_productivity_EGB11 = 0
            Var_productivity_EGB12 = 0
 
 
           
        if 'EGB8' in columna_tareas:
            EGB_Group.append('EGB8')
            last_group = 'EGB8'
            cont_EGB8 = cont_EGB8 + 1
            if columna_horas == 0:
                cont_grupo_EGB8 = cont_grupo_EGB8+ 1
        elif 'EGB9' in columna_tareas:
            EGB_Group.append('EGB9')
            last_group = 'EGB9'
            cont_EGB9 = cont_EGB9 + 1
            if columna_horas == 0:
                cont_grupo_EGB9 = cont_grupo_EGB9+ 1
        elif 'EGB10' in columna_tareas:
            EGB_Group.append('EGB10')
            last_group = 'EGB10'
            cont_EGB10 = cont_EGB10 + 1
            if columna_horas == 0:
                cont_grupo_EGB10 = cont_grupo_EGB10 + 1
        elif 'EGB11' in columna_tareas:
            EGB_Group.append('EGB11')
            last_group = 'EGB11'
            cont_EGB11 = cont_EGB11 + 1
            if columna_horas == 0:
                cont_grupo_EGB11 = cont_grupo_EGB11 + 1
        elif 'EGB12' in columna_tareas:
            EGB_Group.append('EGB12')
            last_group = 'EGB12'
            cont_EGB12 = cont_EGB12 + 1
            if columna_horas == 0:
                cont_grupo_EGB12 = cont_grupo_EGB12 + 1
        else:
            EGB_Group.append(' ')
 
    elif columna_tareas.startswith("   P_") and conditional == 1 and columna_horas != 0:
        Nombres_Tareas.append(columna_tareas)
        hours.append(columna_horas)
        EGB_Group.append(' ')
        Productivity_sum = Productivity_sum + columna_horas
        Var_productivity = Var_productivity + columna_horas
        if last_group == 'EGB8':
            Productivity_abs_EGB8 = Productivity_abs_EGB8 + columna_horas
            Var_productivity_EGB8 = Var_productivity_EGB8 + columna_horas
        elif last_group == 'EGB9':
            Productivity_abs_EGB9 = Productivity_abs_EGB9 + columna_horas
            Var_productivity_EGB9 = Var_productivity_EGB9 + columna_horas
        elif last_group == 'EGB10':
            Productivity_abs_EGB10 = Productivity_abs_EGB10 + columna_horas
            Var_productivity_EGB10 = Var_productivity_EGB10 + columna_horas
        elif last_group == 'EGB11':
            Productivity_abs_EGB11 = Productivity_abs_EGB11 + columna_horas
            Var_productivity_EGB11 = Var_productivity_EGB11 + columna_horas
        elif last_group == 'EGB12':
            Productivity_abs_EGB12 = Productivity_abs_EGB12 + columna_horas
            Var_productivity_EGB12 = Var_productivity_EGB12 + columna_horas
 
   
    elif conditional == 1  and columna_horas != 0 and not 'EGB3' in columna_tareas:
        Nombres_Tareas.append(columna_tareas)
        hours.append(columna_horas)
        EGB_Group.append(' ')
 
   
    elif 'EGB3' in columna_tareas or (columna_tareas.endswith('-MS)') or columna_tareas.endswith('-MX)') or columna_tareas.endswith('-SX)')):
        conditional = 0
        last_group = ' '
   
Productivity.append(Var_productivity/148)
 
People_per_group.append(cont_EGB8-cont_grupo_EGB8)
People_per_group.append(cont_EGB9-cont_grupo_EGB9)
People_per_group.append(cont_EGB10-cont_grupo_EGB10)
People_per_group.append(cont_EGB11-cont_grupo_EGB11)
People_per_group.append(cont_EGB12-cont_grupo_EGB12)
 
 
###########################################################
#                   Archivo Output
# #########################################################
   
 
while True:
    easygui.msgbox("Put the name for you output file")
    archivos = str(easygui.enterbox(msg="Put a name to the output file "))  
    archivo  = archivos + ".xlsx"

    easygui.msgbox("Enter the folder where you want the output file to be stored ")
    directorio=str(easygui.diropenbox())
    mes = nombrecolumna2
    HOURS_PER_DAY = 8.25
    
    ruta_completa = os.path.join(directorio, archivo)
    print(ruta_completa)
    
    # Verificar si el archivo existe
    
    if not os.path.exists(ruta_completa):
        workbook = openpyxl.Workbook()      #Crear un nuevo libro
        hoja = workbook.active              #Crear una nueva hoja
        hoja.title = mes                    #Título de la hoja = mes
        workbook.save(ruta_completa)        #Guardar el libro
    
        easygui.msgbox(f"The file {archivo} was created in: {directorio}")
        #easygui.msgbox("This file was created")
        break
    else:
        #easygui.msgbox("This file already exists")
        easygui.msgbox(f"The file {archivo} already exists in: {directorio}")
        workbook = openpyxl.load_workbook(ruta_completa)
        hoja = workbook.active              #Crear una nueva hoja
        hoja.title = mes                    #Título de la hoja = mes
        workbook.save(ruta_completa)        #Guardar el libro
 
 
#Imprimir encabezados
hoja['A1'] = 'Resource Name'
hoja['B1'] = 'Hours'
hoja['C1'] = 'Productivity'
hoja['D1'] = 'Comments'
hoja['E1'] = 'EGB Group'
hoja['G1'] = 'Days'
hoja['G2'] = WORK_DAYS
hoja['H1'] = 'Hr x day'
hoja['H2'] = HOURS_PER_DAY
hoja['I1'] = 'Total hrs'
TOTAL_HOURS = HOURS_PER_DAY*WORK_DAYS
hoja['I2'] = TOTAL_HOURS
 
Productivity_abs = (Productivity_sum/148) / (cont-cont_grupo)
print(cont_relEGB8,
cont_relEGB9,
cont_relEGB10,
cont_relEGB11,
cont_relEGB12
)

Productivity_rel = (Productivity_sum/148) 

Productivity_rel_EGB8 = (Productivity_abs_EGB8/148) 
Productivity_rel_EGB9 = (Productivity_abs_EGB9/148) 
Productivity_rel_EGB10 = (Productivity_abs_EGB10/148) 
Productivity_rel_EGB11 = (Productivity_abs_EGB11/148) 
Productivity_rel_EGB12 = (Productivity_abs_EGB12/148)

if cont_rel != 0:
    Productivity_rel /= cont_rel

if cont_relEGB8 != 0:
    Productivity_rel_EGB8 /= cont_relEGB8

if cont_relEGB9 != 0:
    Productivity_rel_EGB9 /= cont_relEGB9

if cont_relEGB10 != 0:
    Productivity_rel_EGB10 /= cont_relEGB10

if cont_relEGB11 != 0:
    Productivity_rel_EGB11 /= cont_relEGB11

if cont_relEGB12 != 0:
    Productivity_rel_EGB12 /= cont_relEGB12
    

Productivity_abs_EGB8 = (Productivity_abs_EGB8/148) / (cont_EGB8-cont_grupo_EGB8)
Productivity_abs_EGB9 = (Productivity_abs_EGB9/148) / (cont_EGB9-cont_grupo_EGB9)
Productivity_abs_EGB10 = (Productivity_abs_EGB10/148) / (cont_EGB10-cont_grupo_EGB10)
Productivity_abs_EGB11 = (Productivity_abs_EGB11/148) / (cont_EGB11-cont_grupo_EGB11)
Productivity_abs_EGB12 = (Productivity_abs_EGB12/148) / (cont_EGB12-cont_grupo_EGB12)


 
Productivity_abs_array = []
Productivity_abs_array.append(Productivity_abs_EGB8)
Productivity_abs_array.append(Productivity_abs_EGB9)
Productivity_abs_array.append(Productivity_abs_EGB10)
Productivity_abs_array.append(Productivity_abs_EGB11)
Productivity_abs_array.append(Productivity_abs_EGB12)
 
Productivity_rel_array = []
Productivity_rel_array.append(Productivity_rel_EGB8)
Productivity_rel_array.append(Productivity_rel_EGB9)
Productivity_rel_array.append(Productivity_rel_EGB10)
Productivity_rel_array.append(Productivity_rel_EGB11)
Productivity_rel_array.append(Productivity_rel_EGB12)
 
#Resultados
hoja['H5'] = 'High Value (+10%)'
HIGH_VALUE = TOTAL_HOURS*1.1
hoja['H6'] = HIGH_VALUE
hoja['I5'] = 'Low Value (-5%)'
LOW_VALUE = TOTAL_HOURS*0.95
hoja['I6'] = LOW_VALUE
 
hoja['G9'] = 'High Productivity'
hoja['G10'] = '> 85%'
hoja['H9'] = 'Med Productivity'
hoja['H10'] = '75 % - 85 %'
hoja['I9'] = 'Low Productivity'
hoja['I10'] = '< 75%'
 
hoja['G13'] = 'Qty People'
hoja['G14'] = (cont - cont_grupo)
hoja['H13'] = 'Average Abs. Productivity'
hoja['H14'] = Productivity_abs
hoja['I13'] = 'Average Rel. Productivity'
hoja['I14'] = Productivity_rel
hoja['J13'] = 'Group'
hoja['J14'] = 'BEG'
 
#Imprime productividades absolutas grupales
for i in range (5):
    celda = hoja.cell(row = i+15, column = 8)
    celda.value = Productivity_abs_array[i]
 
#Imprime productividades relativas grupales
for i in range (5):
    celda = hoja.cell(row = i+15, column = 9)
    celda.value = Productivity_rel_array[i]
 
#Imprime personas por equipo
for i in range (5):
    celda = hoja.cell(row = i+15, column = 7)
    celda.value = People_per_group[i]
 
#Imprime equipos
for i in range (5):
    celda = hoja.cell(row = i+15, column = 10, value = Groups[i])
   
 
#Imprimir nombres y tareas
for filas, contenido in enumerate(Nombres_Tareas, start=1):
    hoja.cell(row=filas+1, column=1, value=contenido)
 
#Imprimir horas
for filas, contenido in enumerate(hours, start=1):
    hoja.cell(row=filas+1, column=2, value=contenido)
 
#Imprimir EGB group
for filas, contenido in enumerate(EGB_Group, start=1):
    hoja.cell(row=filas+1, column=5, value=contenido)
 
 
#Imprimir Productividad
counter = 0
for filas, contenido in enumerate(Nombres_Tareas, start=1):
    if 'EGB' in contenido:      
       hoja.cell(row=filas+1, column=3, value=Productivity[counter])      
       counter +=1
     
         
 
####### ESTABLECER ESTILO DE PORCENTAJE EN ALGUNAS CELDAS ##########
# Verificar si el estilo de porcentaje ya existe
porcentaje_style = None
for style in workbook.style_names:
    if style == 'porcentaje_style':
        porcentaje_style = style
        break
 
# Si el estilo no existe, se crea
if porcentaje_style is None:
    porcentaje_style = openpyxl.styles.NamedStyle(name='porcentaje_style')
    porcentaje_style.number_format = '0.00%'
 
columna = 3
filas_1 = 2
 
# Aplicar el estilo a la columna de productividad
for fila in hoja.iter_rows(min_row=2, min_col=columna, max_col=columna, values_only=True):
    celda = hoja.cell(row=filas_1, column=columna)
    celda.style = porcentaje_style
    filas_1 += 1
 
 
 
 
 
####################################################
################# COLOR DE CELDAS ##################
####################################################
 
# Define los colores de relleno
relleno_verde__titulos = PatternFill(start_color="41ad2c", end_color="41ad2c", fill_type="solid")
relleno_verde = PatternFill(start_color="7ce12e", end_color="7ce12e", fill_type="solid")
relleno_amarillo = PatternFill(start_color="d9e221", end_color="d9e221", fill_type="solid")
relleno_rojo = PatternFill(start_color="f25757", end_color="f25757", fill_type="solid")
relleno_gris = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
relleno_naranja = PatternFill(start_color="feb055", end_color="feb055", fill_type="solid")
 
relleno_egb = PatternFill(start_color="ff9e9e", end_color="ff9e9e", fill_type="solid")
relleno_egb8 = PatternFill(start_color="faffa1", end_color="faffa1", fill_type="solid")
relleno_egb9 = PatternFill(start_color="c9ff92", end_color="c9ff92", fill_type="solid")
relleno_egb10 = PatternFill(start_color="92e6ff", end_color="92e6ff", fill_type="solid")
relleno_egb11 = PatternFill(start_color="c77bff", end_color="c77bff", fill_type="solid")
relleno_egb12 = PatternFill(start_color="ff7bd6", end_color="ff7bd6", fill_type="solid")
 
#Defenir el estilo de letras
negritas = Font(bold = True)
letras_rojas = Font(color="FF0000")
 
#### Aplicar los estilos a los Títulos ########
#Encabezados
for i in range (9):
    if i != 5:
        celda = hoja.cell(row = 1, column = i+1)
        celda.fill = relleno_verde__titulos
        celda.font = negritas
 
#Days, hours per days, total hrs
for i in range (3):
    celda = hoja.cell(row = 2, column = i+7)
    celda.fill = relleno_amarillo
 
#High and low hours values headers
for i in range (2):
    celda = hoja.cell(row = 5, column = i+8)
    celda.fill = relleno_verde__titulos
    celda.font = negritas
 
#High Hours Value result
celda = hoja['H6']
celda.fill = relleno_verde
celda.font = negritas
 
#Low value result
celda = hoja['I6']
celda.fill = relleno_rojo
celda.font = negritas
 
 
#High and low Productivity headers
for i in range (3):
    celda = hoja.cell(row = 9, column = i+7)
    celda.fill = relleno_verde__titulos
    celda.font = negritas
 
#High Productivity result
celda = hoja['G10']
celda.fill = relleno_verde
celda.font = negritas
 
#Medium Productivity result
celda = hoja['H10']
celda.fill = relleno_amarillo
celda.font = negritas
 
#Low Productivity result
celda = hoja['I10']
celda.fill = relleno_rojo
celda.font = negritas
 
 
#Productivity Headers
for i in range (4):
    celda = hoja.cell(row = 13, column = i+7)
    celda.fill = relleno_verde__titulos
    celda.font = negritas
 
#Productivity Results
for i in range (6):
    celda = hoja.cell(row = i+14, column = 7)
    celda.fill = relleno_amarillo
    celda.font = negritas
   
 
#Productividad grupal
for i in range (6):
    for j in range(2):
        celda = hoja.cell(row = i+14, column = j+8)
        celda.style = porcentaje_style
        if celda.value is not None:
            if celda.value >= 0.85:
                celda.fill = relleno_verde
            elif celda.value < 0.75:
                celda.fill = relleno_rojo
            else:
                celda.fill = relleno_amarillo
            celda.font = negritas
       
 
   
 
 
 
 
###### Cambiar el formato de nombres de asociados, horas, productividad y grupo ######
for filas, contenido in enumerate(Nombres_Tareas, start=1):
    celda_nombre = hoja.cell(row=filas+1, column=1)
    celda_horas = hoja.cell(row=filas+1, column=2)
    celda_prod = hoja.cell(row=filas+1, column=3)
    celda_egb = hoja.cell(row=filas+1, column=5)
 
 
    if "EGB" in contenido:
        #Aplica formato a los nombres de asociado
        celda_nombre.fill = relleno_gris        
        celda_nombre.font = negritas
        celda_nombre.font = letras_rojas
 
        #Aplica formato a la suma de horas
        if celda_horas.value >= HIGH_VALUE:
            celda_horas.fill = relleno_naranja
        elif celda_horas.value < LOW_VALUE:
            celda_horas.fill = relleno_rojo
        else:
            celda_horas.fill = relleno_verde
        celda_horas.font = negritas
 
        #Aplica formato a las productividades
        if celda_prod.value >= 0.85:
            celda_prod.fill = relleno_verde
        elif celda_prod.value < 0.75:
            celda_prod.fill = relleno_rojo
        else:
            celda_prod.fill = relleno_amarillo
        celda_prod.font = negritas
 
 
        #Aplica formato a la columna de grupos
        if 'EGB8' in celda_egb.value:
            celda_egb.fill = relleno_egb8
            celda_egb.font = negritas
        elif 'EGB9' in celda_egb.value:
            celda_egb.fill = relleno_egb9
            celda_egb.font = negritas
        elif 'EGB10' in celda_egb.value:
            celda_egb.fill = relleno_egb10
            celda_egb.font = negritas
        elif 'EGB11' in celda_egb.value:
            celda_egb.fill = relleno_egb11
            celda_egb.font = negritas
        elif 'EGB12' in celda_egb.value:
            celda_egb.fill = relleno_egb12
            celda_egb.font = negritas
        else:
            celda_egb.fill = relleno_egb
            celda_egb.font = negritas
       
 
#Dar formato a la celdas del grupo del área de productividad
for i in range(6):
    celda = hoja.cell(row = i+14, column = 10)
    if 'EGB8' in celda.value:
        celda.fill = relleno_egb8
    elif 'EGB9' in celda.value:
        celda.fill = relleno_egb9
    elif 'EGB10' in celda.value:
        celda.fill = relleno_egb10
    elif 'EGB11' in celda.value:
        celda.fill = relleno_egb11
    elif 'EGB12' in celda.value:
        celda.fill = relleno_egb12
    else:
        celda.fill = relleno_egb
    celda.font = negritas
 
# CREAR TABLA
existing_tables = hoja.tables
table_name = 'Tabla_general'
 
if table_name in existing_tables:
    # Si la tabla ya existe, la eliminamos
    hoja._tables.remove(existing_tables[table_name])
 
 
#Crea una nueva tabla
table = Table(displayName=table_name, ref=("A1:E")+str(len(Nombres_Tareas)+1))
style = TableStyleInfo(
    name="TableStyleMedium9", showFirstColumn=False,
    showLastColumn=False, showRowStripes=False, showColumnStripes=False)
table.tableStyleInfo = style
hoja.add_table(table)
 
 
 
# Ajustar el ancho de las columnas de la tabla
for columna in hoja.iter_cols(min_col=1, max_col=5):
    longitud_maxima = 0
    columna_letra = columna[0].column_letter  # Obtiene la letra de la columna
    for celda in columna:
        if celda.value:
            # Calcula la longitud máxima del contenido en la columna
            longitud_celda = len(str(celda.value))
            if longitud_celda > longitud_maxima:
                longitud_maxima = longitud_celda
 
    # Establece el ancho de la columna para ajustarlo al contenido más largo
    hoja.column_dimensions[columna_letra].width = longitud_maxima + 4
 
# Ajustar el ancho de las columnas de los resultados
for columna in hoja.iter_cols(min_col=7, max_col=9):
    longitud_maxima = 0
    columna_letra = columna[0].column_letter  # Obtiene la letra de la columna
    for celda in columna:
        if celda.value:
            # Calcula la longitud máxima del contenido en la columna
            longitud_celda = len(str(celda.value))
            if longitud_celda > longitud_maxima:
                longitud_maxima = longitud_celda
 
    # Establece el ancho de la columna para ajustarlo al contenido más largo
    hoja.column_dimensions[columna_letra].width = longitud_maxima -2
 
 
workbook.save(ruta_completa)
 
if os.path.exists(ruta_completa):
    easygui.msgbox("Save completed, opening file...")
    os.startfile(ruta_completa)
    #os.system(f'start excel "{ruta_completa}"')
else:
    easygui.msgbox(f'El archivo "{ruta_completa}" no existe.')
 
