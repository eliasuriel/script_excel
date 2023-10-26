import pandas as pd
import numpy as np
import os
import openpyxl
import easygui
from openpyxl.styles import PatternFill
from openpyxl.styles import NamedStyle

import tkinter as tk
from tkinter import filedialog

root = tk.Tk()
root.withdraw()

file_path = filedialog.askopenfilename()


#df = pd.read_excel("Script_Input.xlsx") 

# Reemplaza 'nombre_del_archivo.xlsx' con el nombre de tu archivo Excel
excel_file = pd.ExcelFile(file_path)
easygui.msgbox("The file was processed successfully")
hojas = str(easygui.enterbox(msg="In which sheet do you want to work? "))

#Hojas es la hojas donde se trabjara
#Sheet1
df = excel_file.parse(hojas)

nombrecolumna1 = str(easygui.enterbox(msg="Put the name of the first column you want to work on (coworkers,name of the jobs) "))
nombrecolumna2 = str(easygui.enterbox(msg="Put the name of the second column you want to work on (hours) "))

# Reemplaza los valores nulos en la columna 2 con ceros
df[nombrecolumna2] = df[nombrecolumna2].fillna(0)
columna2 = df[nombrecolumna2]
columna1 = df[nombrecolumna1]
df['EGB_Group'] = '' 


Nombres_Tareas = []
hours = []
Productivity = []
EGB_Group = []

cont = 0
conditional = 0
Productivity_abs = 0
Productivity_rel = 0
Productivity_sum = 0
Var_productivity = 0
cont_rel = 0

WORK_DAYS = int(easygui.enterbox(msg="How many work days are in this report?: "))
easygui.msgbox(WORK_DAYS, title="Días de trabajo ")

for index, row in df.iterrows():
    columna_tareas = row[nombrecolumna1]
    columna_horas = row[nombrecolumna2]
    

    if 'EGB' in columna_tareas:
        cont = cont + 1
        Nombres_Tareas.append(columna_tareas)
        hours.append(columna_horas)
        conditional = 1

        if cont > 1:
            Productivity.append(Var_productivity/148)

            if Var_productivity != 0:
                cont_rel = cont_rel + 1

            Var_productivity = 0

            
        if 'EGB3' in columna_tareas:
            EGB_Group.append('EGB3')
        elif 'EGB8' in columna_tareas:
            EGB_Group.append('EGB8')
        elif 'EGB9' in columna_tareas:
            EGB_Group.append('EGB9')
        elif 'EGB10' in columna_tareas:
            EGB_Group.append('EGB10')
        elif 'EGB11' in columna_tareas:
            EGB_Group.append('EGB11')
        elif 'EGB12' in columna_tareas:
            EGB_Group.append('EGB12')
        else:
            EGB_Group.append(' ')

    elif columna_tareas.startswith("   P_") and conditional == 1 and columna_horas != 0:
        Nombres_Tareas.append(columna_tareas)
        hours.append(columna_horas)
        EGB_Group.append(' ')
        Productivity_sum = Productivity_sum + columna_horas
        Var_productivity = Var_productivity + columna_horas

    
    elif conditional == 1  and columna_horas != 0:
        Nombres_Tareas.append(columna_tareas)
        hours.append(columna_horas)
        EGB_Group.append(' ')

    
    elif not 'EGB' in columna_tareas and (columna_tareas.endswith('-MS)') or columna_tareas.endswith('-MX)') or columna_tareas.endswith('-SX)')):
        conditional = 0
    
Productivity.append(Var_productivity/148)




###########################################################
#                   Archivo Output
# #########################################################
easygui.msgbox("Add .xlsx to the output file name")
archivo = str(easygui.enterbox(msg="Put a name to the output file "))  

#archivo = "Monthly_Output.xlsx"
directorio = "C:\\Users\\VEU1GA\\Documents\\Visual_Studio"
mes = "Sep"
HOURS_PER_DAY = 8.25

ruta_completa = os.path.join(directorio, archivo)
print(ruta_completa)

# Verificar si el archivo existe

if not os.path.exists(ruta_completa):
    workbook = openpyxl.Workbook()      #Crear un nuevo libro
    hoja = workbook.active              #Crear una nueva hoja
    hoja.title = mes                    #Título de la hoja = mes
    workbook.save(ruta_completa)        #Guardar el libro

    print(f"Se ha creado el archivo {archivo} en: {ruta_completa}")

else:
    print(f"El archivo {archivo} ya existe en: {ruta_completa}")
    workbook = openpyxl.load_workbook(ruta_completa)
    hoja = workbook.active              #Crear una nueva hoja
    hoja.title = mes                    #Título de la hoja = mes
    workbook.save(ruta_completa)        #Guardar el libro

#Imprimir encabezados
hoja['A1'] = 'Resource Name'
hoja['B1'] = 'Hours'
hoja['C1'] = 'Productivity'
hoja['D1'] = 'Comments'
hoja['E1'] = 'EGB Group'
hoja['G1'] = 'Days'
hoja['G2'] = WORK_DAYS
hoja['H1'] = 'Hr x day'
hoja['H2'] = HOURS_PER_DAY
hoja['I1'] = 'Total hrs'
hoja['I2'] = HOURS_PER_DAY*WORK_DAYS

Productivity_abs = (Productivity_sum/148) / cont
Productivity_rel = (Productivity_sum/148) / cont_rel


#Resultados
hoja['G5'] = 'Abs. Productivity Sum(%)'
hoja['G6'] = Productivity_sum/148
hoja['H5'] = 'Qty People'
hoja['H6'] = cont
hoja['I5'] = 'Average Abs. Productivity'
hoja['I6'] = Productivity_abs
hoja['J5'] = 'Average Rel. Productivity'
hoja['J6'] = Productivity_rel


#Imprimir nombres y tareas
for filas, contenido in enumerate(Nombres_Tareas, start=1):
    hoja.cell(row=filas+1, column=1, value=contenido)

#Imprimir nombres y tareas
for filas, contenido in enumerate(hours, start=1):
    hoja.cell(row=filas+1, column=2, value=contenido)

#Imprimir EGB group
for filas, contenido in enumerate(EGB_Group, start=1):
    hoja.cell(row=filas+1, column=5, value=contenido)


#Imprimir Productividad
counter = 0
for filas, contenido in enumerate(Nombres_Tareas, start=1):
    if 'EGB' in contenido:
       hoja.cell(row=filas+1, column=3, value=Productivity[counter])
       counter +=1  
    

# Itera a través de todas las columnas en la hoja para ajustar su ancho
for columna in hoja.columns:
    longitud_maxima = 0
    columna_letra = columna[0].column_letter  # Obtiene la letra de la columna
    for celda in columna:
        if celda.value:
            # Calcula la longitud máxima del contenido en la columna
            longitud_celda = len(str(celda.value))
            if longitud_celda > longitud_maxima:
                longitud_maxima = longitud_celda

    # Establece el ancho de la columna para ajustarlo al contenido más largo
    hoja.column_dimensions[columna_letra].width = longitud_maxima + 2




# Verificar si el estilo de porcentaje ya existe
porcentaje_style = None
for style in workbook.style_names:
    if style == 'porcentaje_style':
        porcentaje_style = style
        break

# Si el estilo no existe, se crea
if porcentaje_style is None:
    porcentaje_style = openpyxl.styles.NamedStyle(name='porcentaje_style')
    porcentaje_style.number_format = '0.00%'

columna = 3
filas_1 = 2

# Aplicar el estilo a la columna de productividad
for fila in hoja.iter_rows(min_row=2, min_col=columna, max_col=columna, values_only=True):
    celda = hoja.cell(row=filas_1, column=columna)
    celda.style = porcentaje_style
    filas_1 += 1

celda = hoja['G6'] 
celda.style = porcentaje_style

celda = hoja['I6'] 
celda.style = porcentaje_style

celda = hoja ['J6']
celda.style = porcentaje_style

workbook.save(ruta_completa)







#print(Nombres_Tareas)  
#df = df[df['Sep'] != 0]


        

#print(cont)
#print(horas)
#print("OTRA \n" )
#print(horas2)
#print(df)