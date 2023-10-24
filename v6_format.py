import os
import openpyxl
from openpyxl.styles import PatternFill

archivo = "Monthly_Output.xlsx"
directorio = "C:\\Users\OJA5GA\Documents\VS_Code\V1"
mes = "Sep"
WORK_DAYS = 22
HOURS_PER_DAY = 8.25

ruta_completa = os.path.join(directorio, archivo)
print(ruta_completa)

# Verificar si el archivo existe

if not os.path.exists(ruta_completa):
    workbook = openpyxl.Workbook()      #Crear un nuevo libro
    hoja = workbook.active              #Crear una nueva hoja
    hoja.title = mes                    #Título de la hoja = mes
    workbook.save(ruta_completa)        #Guardar el libro

    print(f"Se ha creado el archivo {archivo} en: {ruta_completa}")

else:
    print(f"El archivo {archivo} ya existe en: {ruta_completa}")
    workbook = openpyxl.load_workbook(ruta_completa)
    hoja = workbook.active              #Crear una nueva hoja
    hoja.title = mes                    #Título de la hoja = mes
    workbook.save(ruta_completa)        #Guardar el libro

#Imprimir encabezados
hoja['A1'] = 'Resource Name'
hoja['B1'] = 'Hours'
hoja['C1'] = 'Productivity'
hoja['D1'] = 'Comments'
hoja['E1'] = 'EGB Group'
hoja['G1'] = 'Days'
hoja['G2'] = WORK_DAYS
hoja['H1'] = 'Hr x day'
hoja['H2'] = HOURS_PER_DAY
hoja['I1'] = 'Total hrs'
hoja['I2'] = HOURS_PER_DAY*WORK_DAYS

#Resultados
hoja['G5'] = 'Sum(%)'
hoja['G6'] = '=SUMA(C:C)'
hoja['H5'] = 'Qty People'
hoja['H6'] = '=CONTAR(C:C)'
hoja['I5'] = '% total'
hoja['I6'] = '=G6/H6'

#Imprimir datos a partir de la fila 2
#Imprimir personas y tareas en columna A
#Imprimir horas en columna B
#imprimir EGB group en columna E

#Dar formato de color a un rango de celdas

""" # Define los colores de relleno
relleno_verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
relleno_rojo = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Itera a través de las celdas de la columna A
for fila in hoja.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
    valor_celda = fila[0]
    celda = hoja.cell(row=fila[0].row, column=1)

    # Aplica el formato condicional basado en el valor de la celda
    if valor_celda > 5:
        celda.fill = relleno_verde
    else:
        celda.fill = relleno_rojo """




#Establecer formato de ciertas celdas como porcentaje
""" # Definir un estilo de celda con formato de porcentaje
porcentaje_style = NamedStyle(name='porcentaje_style')
porcentaje_style.number_format = '0.00%'  # Establece el formato de porcentaje

# Aplica el estilo a una celda específica (por ejemplo, A1)
celda = hoja['A1']
celda.value = 0.256  # Este valor se mostrará como 25.60%
celda.style = porcentaje_style """





# Itera a través de todas las columnas en la hoja para ajustar su ancho
for columna in hoja.columns:
    longitud_maxima = 0
    columna_letra = columna[0].column_letter  # Obtiene la letra de la columna
    for celda in columna:
        if celda.value:
            # Calcula la longitud máxima del contenido en la columna
            longitud_celda = len(str(celda.value))
            if longitud_celda > longitud_maxima:
                longitud_maxima = longitud_celda

    # Establece el ancho de la columna para ajustarlo al contenido más largo
    hoja.column_dimensions[columna_letra].width = longitud_maxima + 5

workbook.save(ruta_completa)





