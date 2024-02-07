# Importar las bibliotecas necesarias
import pandas as pd
import numpy as np
import os
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import NamedStyle
from openpyxl.worksheet.table import Table, TableStyleInfo

import tkinter as tk
from tkinter import filedialog
import easygui
import random

###############################################
#  1. Importacion y modificacion de archivos  #
###############################################


# Mostrar un cuadro de mensaje de bienvenida
easygui.msgbox("Bienvenido. Elija su archivo de Excel con datos de PRIME")

# Crear una ventana de selección de archivo
root = tk.Tk()
root.withdraw()


file_path = filedialog.askopenfilename()

# Pedir al usuario que elija el archivo de Excel con datos de OPX
easygui.msgbox("Elija su archivo de Excel con datos de OPX")
file_path_opx = filedialog.askopenfilename()

# Crear objetos de archivo Excel a partir de los archivos seleccionados
excel_file = pd.ExcelFile(file_path)
opx_file = pd.ExcelFile(file_path_opx)

# Realizar cambios en el archivo OPX para celdas combinadas
hoja_nombre = opx_file.sheet_names[0]

# Obtener el DataFrame de la hoja especificada en el archivo OPX
df_2 = opx_file.parse(hoja_nombre)

# Crear un objeto Workbook de openpyxl a partir del DataFrame de OPX
workbook_de_paso = openpyxl.Workbook()
hoja_prueba = workbook_de_paso.active

# Copiar los datos del DataFrame a la hoja de openpyxl
for i, row in enumerate(df_2.values):
    for j, value in enumerate(row):
        hoja_prueba.cell(row=i+1, column=j+1, value=value)

# Descombinar las celdas en la hoja
for rango in hoja_prueba.merged_cells.ranges:
    # Obtener el valor de la celda combinada
    valor_celda_combinada = hoja_prueba[rango.start_cell].value
    
    # Iterar sobre cada celda en el rango y asignarle el mismo valor
    for celda in rango.cells:
        hoja_prueba[celda].value = valor_celda_combinada

# Copiar valores de las primeras dos celdas a las nuevas celdas de la segunda fila
celda1 = hoja_prueba['A1']
celda2 = hoja_prueba['B1']
hoja_prueba.cell(row=2, column=1, value=celda1.value)
hoja_prueba.cell(row=2, column=2, value=celda2.value)

# Eliminar la primera fila que ahora está vacía
hoja_prueba.delete_rows(1)

# Obtener el valor de la celda en la columna C de la primera fila, contiene el primer anio
celda1 = hoja_prueba['C1']

# Verificar si el valor de la celda contiene '202' porque significa que hay un anio
if '202' in celda1.value:
    for i in range(11):
        # Asigna a las celdas despues del anio, su valor mas su identificador decimal
        celda = hoja_prueba.cell(row=1, column=4+i, value=celda1.value +"."+str(i+1))

# Obtener el valor de la celda en la columna O de la primera fila, que contiene el 2do anio
celda2 = hoja_prueba['O1']

# Verificar si el valor de la celda2 contiene '202' porque significa que hay un anio
if '202' in celda2.value:
    for i in range(11):
        # Asigna a las celdas despues del anio, su valor mas su identificador decimal
        celda = hoja_prueba.cell(row=1, column=16+i, value=celda2.value +"."+str(i+1))


#### Nombre del archivo de SALIDA ###
num_random = random.randint(1,999)
archivos = "Output-" + str(num_random)
#archivos = str(easygui.enterbox(msg="Escribe un nombre para el archivo de salida"))
archivo  = archivos + ".xlsx"

# Solicitar al usuario la carpeta donde desea guardar el archivo de salida
easygui.msgbox("Ingresa la carpeta donde deseas guardar el archivo de salida")
directorio = str(easygui.diropenbox())

# Crear la ruta completa para el archivo de salida
ruta_completa = os.path.join(directorio, archivo)
print(ruta_completa)


# Solicitar al usuario un nombre para el archivo modificado OPX
#archivos_temp = str(easygui.enterbox(msg="Escribe un nombre para el archivo OPX modificado"))

archivos_temp = "OPX_modified-" + str(num_random)
archivo_temp  = archivos_temp + ".xlsx"
 
ruta_completa_temp = os.path.join(directorio, archivo_temp)

# Guardar el Workbook modificado en la nueva ruta
workbook_de_paso.save(ruta_completa_temp)

# Crear un nuevo objeto ExcelFile a partir del archivo modificado OPX
opx_file = pd.ExcelFile(ruta_completa_temp)

# Continuar con el proceso
easygui.msgbox("Los archivos fueron procesados exitosamente")



###################################################
#       2. Definicion de parametros               #
###################################################
# Solicitar al usuario ingresar el anio en el que desea trabajar
anio =  int(easygui.enterbox("Ingresa el año en el que deseas trabajar"))
str_anio = str(anio)

numhojas = 0

# Listas para almacenar información sobre hojas y columnas
hojas = []
hojas_2 = []

# Obtener los nombres de las hojas del archivo OPX
hojas_2 = opx_file.sheet_names

# Nombres de las columnas del archivo PRIME
nombrecolumna1 = "Resource Name"
nombrecolumna2 = "Hours"

# Listas para almacenar información sobre tareas, horas y proyectos
Nombres_Tareas = []
hours = []
column_headers = []
seleccion_hoja = []
Proyectos = []

# Inicializar variables de control y acumuladores
cont = 0
cont_anio = 0
suma_prime = 0
anio_mes = ""
palabra = "BE-"
palabra2 = "PN-"
condicional = 0
suma_opx = 0
totalopx = 0
mesopx = ""
res_prime = []
res_opx = []
Nombres = ''


########################################
#   3. Creacion Archivo Output         #
########################################
# Solicitar al usuario un nombre para el archivo de salida


# Verificar si el archivo de salida ya existe
if not os.path.exists(ruta_completa):
    # Si no existe, crear un nuevo libro de Excel y hojas
    workbook = openpyxl.Workbook()      
    hoja = workbook.active            
    hoja.title = str_anio               
    hoja2 = workbook.create_sheet(title="Lista asociados " + str_anio)             
    workbook.save(ruta_completa)        
    easygui.msgbox(f"El archivo {archivo} fue creado en: {directorio}")
else:
    # Si ya existe, cargar el libro existente y crear hojas
    easygui.msgbox(f"El archivo {archivo} ya existe en: {directorio}")
    workbook = openpyxl.load_workbook(ruta_completa)
    hoja = workbook.create_sheet              
    hoja.title = str_anio               
    hoja2 = workbook.create_sheet             
    hoja2.title = "Lista asociados " + str_anio
    workbook.save(ruta_completa)



#################################################
# 4. Manejo de datos e impresion en output      #
#################################################
    
# Solicitar al usuario ingresar la cantidad de horas laborales que tiene un día
HOURS_PER_DAY = float(easygui.enterbox("¿Cuántas horas laborales tiene un día? (8, 8.25...)"))

# Inicializar variables para filas y columnas
num_fila = 1
num_columna = 1

# Definir estilos de celdas
relleno_titulos = PatternFill(start_color="77bedc", end_color="77bedc", fill_type="solid")
negritas = Font(bold=True)

# Agregar títulos a las celdas B1 y C1
celda = hoja["B1"]
celda.value = "Hours/day"
celda.fill = relleno_titulos
celda.font = negritas

celda = hoja["C1"]
celda.value = HOURS_PER_DAY
celda.fill = relleno_titulos
celda.font = negritas

# Definir componentes y proyectos
area = ["CSW", "NET", "DCOM", "DSW", "TST", "SYS", "ASW"]
proyectos = ["BRP", "Zoox", "Faraday", "Tesla", "Ford", "Singer", "Harley Davidson", "GM", "Lordstwon", "Oshkosh", "Lucid", "Navistar", "Rexroth", "METALSA", "ASNA", "ReCar", "Fisker", "Battle_Motors", "BYD"]

# Definir personas excluidas
gente_sin_egb = [
    "Humberto Chávez Chávez",
    "Ricardo Adolfo del Razo Real",
    "Atlahuac",
    "Aguirre Vivo Mayra Fernanda",
    "Diaz Paredes Jose Ronaldo",
    "Perez Mejia Marcelo Armando",
    "Dominguez Barba Jose Ramon",
    "Centeno Alcaraz Jaime Alberto",
    "Perez Fonseca Erick Zahid"
]

# Solicitar al usuario ingresar la cantidad de proyectos en los que desea trabajar
#num_proyectos = int(easygui.enterbox("¿Cuántos proyectos deseas incluir en el trabajo?"))

# Inicializar variables para filas y columnas de la segunda hoja, la de asociados
num_columna_2 = 2
num_fila_2 = 2

# Agregar títulos a las celdas de la hoja de asociados
celda = hoja2.cell(row=num_fila_2, column=num_columna_2, value="Project")
celda.fill = relleno_titulos
celda = hoja2.cell(row=num_fila_2, column=num_columna_2 + 1, value="Month")
celda.fill = relleno_titulos
celda = hoja2.cell(row=num_fila_2, column=num_columna_2 + 2, value="Component")
celda.fill = relleno_titulos
celda = hoja2.cell(row=num_fila_2, column=num_columna_2 + 3, value="Name")
celda.fill = relleno_titulos
celda = hoja2.cell(row=num_fila_2, column=num_columna_2 + 4, value="Activity")
celda.fill = relleno_titulos
celda = hoja2.cell(row=num_fila_2, column=num_columna_2 + 5, value="Hours Reported")
celda.fill = relleno_titulos
num_fila_2 += 1



#######################################
#      5. Iteraciones y Calculos      #
#######################################
num_proyectos = 0

#Arreglo que contiene los nombres de las hojas del PRIME
meses = [
    "Associate_Jan",
    "Associate_Feb",
    "Associate_March",
    "Associate_April",
    "Associate_May",
    "Associate_June",
    "Associate_July",
    "Associate_Aug",
    "Associate_Sep",
    "Associate_Oct",
    "Associate_Nov",
    "Associate_Dec"
]

#Se le pide al usuario que seleccione los proyectos
seleccion_proyecto = easygui.multchoicebox(f"Selecciona en qué proyectos deseas trabajar?","Proyectos",proyectos)
num_proyectos = len(seleccion_proyecto)

# Obtener los nombres de las hojas (o meses) del archivo Prime
nombres_hojas_prime = excel_file.sheet_names


while True:
    seleccion_hoja = []
    numhojas = 0
    contnum = 0

    seleccion_hoja = easygui.multchoicebox(f"Selecciona en qué meses del archivo Prime deseas trabajar? ","Meses",meses)
    numhojas = len(seleccion_hoja)
    
    for a in range(numhojas):
        if str(seleccion_hoja[a]) in nombres_hojas_prime:
            contnum += 1
        else:
            easygui.msgbox("El mes " +str(seleccion_hoja[a]) + " no se encontró en archivo PRIME", "ERROR")
    if contnum == numhojas:
        break

   
# Iterar sobre el número de proyectos ingresado por el usuario
for j in range(num_proyectos):
    
    # Inicializar listas para almacenar meses y hojas
    mes = []
    hojas = []

    # Solicitar al usuario seleccionar un proyecto de la lista
    #seleccion_proyecto = easygui.choicebox(f"Elige el {j + 1}° proyecto en el que deseas trabajar", choices=proyectos)
    proyecto = seleccion_proyecto[j]

    

    # Iterar sobre el número de meses ingresado por el usuario
    for i in range(numhojas):
        # Solicitar al usuario seleccionar una hoja (o mes) del archivo Prime
        #seleccion = easygui.choicebox(f"¿En qué hoja (o mes) del archivo Prime deseas trabajar? {i + 1}", choices=nombres_hojas_prime)
        hojas.append(seleccion_hoja[i])

        for b in range(len(meses)):
            if seleccion_hoja[i] in meses:
                mes.append(int(b+1))

      

    # Imprimir el proyecto en la hoja principal
    num_columna = 2
    num_fila += 2
    celda = hoja.cell(row=num_fila, column=num_columna, value=proyecto)
    celda.fill = relleno_titulos
    celda.font = negritas

    # Imprimir componentes en la hoja principal
    for h in range(len(area)):
        for k in range(3):
            # Imprimir el nombre del área en las celdas correspondientes
            celda = hoja.cell(row=num_fila, column=(1+k+num_columna+3*h), value=area[h])
            celda.fill = relleno_titulos

        # Imprimir títulos "OPX", "Prime" y "Relation" en las celdas correspondientes
        celda = hoja.cell(row=num_fila+1, column=(1+num_columna+3*h), value="OPX")
        celda.fill = relleno_titulos

        celda = hoja.cell(row=num_fila+1, column=(2+num_columna+3*h), value="Prime")
        celda.fill = relleno_titulos

        celda = hoja.cell(row=num_fila+1, column=(3+num_columna+3*h), value="Relation")
        celda.fill = relleno_titulos

    # Incrementar el número de fila para la siguiente iteración
    num_fila += 2

    # Iterar sobre el número de hojas (o meses) seleccionados para el proyecto
    for i in range(numhojas):
        # Obtener datos de las hojas correspondientes en los archivos Prime y OPX
        df = excel_file.parse(hojas[i])
        df_2 = opx_file.parse(hojas_2[0])

        # Eliminar filas con valores nulos en las columnas especificadas
        df = df.dropna(subset=[nombrecolumna1, nombrecolumna2])
        column_headers = df_2.columns.tolist()
        df_2[column_headers] = df_2[column_headers].fillna(0)

        # Asignar el año y mes correspondiente en el formato deseado
        if mes[i] == 1:
            anio_mes = str_anio 
        elif mes[i] == 2:
            anio_mes = str_anio + ".1"
        elif mes[i] == 3:
            anio_mes = str_anio + ".2"
        elif mes[i] == 4:
            anio_mes = str_anio + ".3"
        elif mes[i] == 5:
            anio_mes = str_anio + ".4"
        elif mes[i] == 6:
            anio_mes = str_anio + ".5"
        elif mes[i] == 7:
            anio_mes = str_anio + ".6"
        elif mes[i] == 8:
            anio_mes = str_anio + ".7"
        elif mes[i] == 9:
            anio_mes = str_anio + ".8"
        elif mes[i] == 10:
            anio_mes = str_anio + ".9"
        elif mes[i] == 11:
            anio_mes = str_anio + ".10"
        elif mes[i] == 12:
            anio_mes = str_anio + ".11"
        else:
            anio_mes = "0"


        # Iterar sobre las áreas
        for h in range(len(area)):
            
            # Iterar sobre los headers de las columnas en el archivo OPX
            for col_header in column_headers:
                
                # Verificar si el año y mes corresponden al header actual
                if str_anio in str(col_header) and anio_mes == str(col_header):
                    
                    # Iterar sobre las filas del DataFrame de Excel
                    for index, row in df.iterrows():
                        columna_tareas = row[nombrecolumna1]
                        columna_horas = row[nombrecolumna2]    
                        
                        # Verificar si la tarea pertenece a personas excluidas
                        if ('EGB' in columna_tareas) or (columna_tareas in gente_sin_egb):
                            Nombres = columna_tareas
                                    
                        # Verificar si la tarea pertenece al proyecto y área especificados
                        if proyecto in str(columna_tareas) and area[h] in str(columna_tareas):
                            suma_prime = suma_prime + columna_horas
                            Nombres_Tareas.append(Nombres)
                            
                            # Agregar información a la hoja secundaria
                            celda = hoja2.cell(row=num_fila_2, column=num_columna_2, value=proyecto)
                            celda = hoja2.cell(row=num_fila_2, column=num_columna_2+1, value=hojas[i])
                            celda = hoja2.cell(row=num_fila_2, column=num_columna_2+2, value=area[h])
                            celda = hoja2.cell(row=num_fila_2, column=num_columna_2+3, value=Nombres)
                            celda = hoja2.cell(row=num_fila_2, column=num_columna_2+4, value=columna_tareas)
                            celda = hoja2.cell(row=num_fila_2, column=num_columna_2+5, value=columna_horas)
                            
                            num_fila_2 += 1
                            
                    # Iterar sobre las filas del DataFrame de OPX
                    for index, row in df_2.iterrows():                    
                        columna_1 = row[column_headers[0]]
                        columna_2 = row[col_header] 
                        
                        # Verificar condiciones y acumular horas OPX
                        if index == 0:
                            mesopx = columna_2
                            celda = hoja.cell(row=num_fila, column=num_columna, value=mesopx)
                            celda.fill = relleno_titulos                     
                        
                        if str(columna_1).startswith(palabra or palabra2):
                            condicional = 0
                                    
                        if proyecto in str(columna_1):
                            condicional = 1
                                    
                        if condicional == 1 and area[h] in str(columna_1):
                            suma_opx = suma_opx + columna_2 * HOURS_PER_DAY


            #easygui.msgbox("Mes Prime: " + str(hojas[i]) + "\nMes OPX: " + str(mesopx) + "\nProyecto: " + str(proyecto) + "\nArea: " + str(area[h])  + "\nHoras Prime: " + str(suma_prime) + "\nHojas OPX: " + str(suma_opx))
        
            #Inicializar lista para almacenar resultados
            resultados = []

            # Agregar la suma de horas Prime y OPX a las listas correspondientes
            res_prime.append(suma_prime)
            res_opx.append(suma_opx) 

            # Agregar las sumas a la lista de resultados
            resultados.append(suma_opx)
            resultados.append(suma_prime)

            # Calcular la relación entre las sumas de Prime y OPX
            if suma_opx == 0:
                relacion = 0
            else:
                relacion = suma_prime / suma_opx
            resultados.append(relacion)

            # Imprimir los resultados en las celdas correspondientes en la hoja principal
            for k in range(3):
                celda = hoja.cell(row=num_fila, column=(1+k+num_columna+3*h), value=resultados[k])

            # Reiniciar las sumas si no son cero
            if suma_prime != 0:
                suma_prime = 0
            if suma_opx != 0:
                suma_opx = 0

                # Incrementar el número de fila para la siguiente iteración
        num_fila += 1

##############################################
#       6. Dar formato a la Hoja             #
##############################################

# Hacer una tabla en la hoja secundaria
existing_tables = hoja2.tables
table_name = 'Tabla_general'
 
# Verificar si la tabla ya existe y eliminarla si es necesario
if table_name in existing_tables:
    hoja2._tables.remove(existing_tables[table_name])

# Crear una nueva tabla
table = Table(displayName=table_name, ref=("B2:G")+str(num_fila_2 - 1))
style = TableStyleInfo(
    name="TableStyleMedium9", showFirstColumn=False,
    showLastColumn=False, showRowStripes=False, showColumnStripes=True)
table.tableStyleInfo = style
hoja2.add_table(table)


# Ajustar el ancho de las columnas en la hoja principal
for columna in hoja.iter_cols(min_col=2, max_col=2+(len(area)*3)):
    longitud_maxima = 0
    columna_letra = columna[0].column_letter  # Obtener la letra de la columna
    for celda in columna:
        if celda.value:
            if celda.data_type == 'n':
                longitud_celda = 8
            else:
                # Calcular la longitud máxima del contenido en la columna
                longitud_celda = len(str(celda.value))
            
            # Actualizar la longitud máxima si es necesario
            if longitud_celda > longitud_maxima:
                longitud_maxima = longitud_celda
 
    # Establecer el ancho de la columna para ajustarlo al contenido más largo
    hoja.column_dimensions[columna_letra].width = longitud_maxima + 2


# Ajustar el ancho de las columnas en la hoja secundaria
for columna in hoja2.iter_cols(min_col=2, max_col=7):
    longitud_maxima = 0
    columna_letra = columna[0].column_letter  # Obtener la letra de la columna
    for celda in columna:
        if celda.value:
            if celda.data_type == 'n':
                longitud_celda = 8
            else:
                # Calcular la longitud máxima del contenido en la columna
                longitud_celda = len(str(celda.value))
            
            # Actualizar la longitud máxima si es necesario
            if longitud_celda > longitud_maxima:
                longitud_maxima = longitud_celda
 
    # Establecer el ancho de la columna para ajustarlo al contenido más largo
    hoja2.column_dimensions[columna_letra].width = longitud_maxima + 4

#######################################
#       7. Ajustes finales            #
#######################################

# Guardar el archivo de salida
workbook.save(ruta_completa)

# Eliminar el OPX modificado
workbook_de_paso.close()
opx_file.close()
if os.path.exists(ruta_completa_temp):
    # Borrar el archivo
    os.remove(ruta_completa_temp) 
 
# Abrir el archivo de salida si existe
if os.path.exists(ruta_completa):
    easygui.msgbox("Save completed, opening file...")
    os.startfile(ruta_completa)
else:
    easygui.msgbox(f'The file "{ruta_completa}" does not exist.')
